#!/usr/bin/env python3
"""Converte um relatório do SGP (xlsx) com colunas de contrato e vendedor em JSON.
Detecta automaticamente as colunas cujo cabeçalho contenha 'contrato' e
'vendedor'/'usuário'. Uso:
  python3 scripts/importar-vendedores.py <arquivo.xlsx> > /tmp/vendedores.json"""
import json, sys, unicodedata
import openpyxl

def norm(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "")) if unicodedata.category(c) != "Mn").lower()

wb = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
linhas = list(ws.iter_rows(values_only=True))

col_contrato = col_vendedor = cab_idx = None
for i, linha in enumerate(linhas[:15]):
    cabs = [norm(c) for c in (linha or [])]
    for j, cab in enumerate(cabs):
        if "contrato" in cab and col_contrato is None:
            col_contrato = j
        if ("vendedor" in cab or "usuario" in cab) and col_vendedor is None:
            col_vendedor = j
    if col_contrato is not None and col_vendedor is not None:
        cab_idx = i
        break

if cab_idx is None:
    print("Não achei colunas de contrato + vendedor no cabeçalho.", file=sys.stderr)
    sys.exit(1)

saida = []
for linha in linhas[cab_idx + 1:]:
    if not linha or linha[col_contrato] in (None, ""):
        continue
    contrato = str(linha[col_contrato]).strip().split(".")[0]
    vendedor = str(linha[col_vendedor] or "").strip()
    if contrato and vendedor:
        saida.append({"contrato": contrato, "vendedor": vendedor})
json.dump(saida, sys.stdout, ensure_ascii=False)
print(f"\n{len(saida)} linhas (colunas: contrato={col_contrato}, vendedor={col_vendedor})", file=sys.stderr)
