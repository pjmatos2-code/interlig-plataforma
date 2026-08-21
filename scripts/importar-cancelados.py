#!/usr/bin/env python3
"""Converte o relatório 'Contratos Cancelados' do SGP (xlsx) em JSON.
Uso: python3 scripts/importar-cancelados.py <arquivo.xlsx> > /tmp/cancelados.json"""
import json, sys
import openpyxl

wb = openpyxl.load_workbook(sys.argv[1], read_only=True)
ws = wb[wb.sheetnames[0]]
saida = []
for linha in ws.iter_rows(values_only=True):
    if not linha or not linha[0] or str(linha[0]).strip() in ("", "Contrato ID"):
        continue
    contrato, _cliente, data, motivo = linha[0], linha[1], linha[2], linha[3]
    d = str(data).strip()[:10]
    if "/" in d:
        dia, mes, ano = d.split("/")
        d = f"{ano}-{mes}-{dia}"
    saida.append({
        "contrato": str(contrato).strip(),
        "data": d,
        "motivo": (str(motivo).replace("Cancelamento - ", "").strip() or None) if motivo else None,
    })
json.dump(saida, sys.stdout, ensure_ascii=False)
print(f"\n{len(saida)} registros", file=sys.stderr)
